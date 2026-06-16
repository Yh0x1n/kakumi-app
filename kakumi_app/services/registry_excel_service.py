"""Shared Excel workbook contract for registry import/export."""

from __future__ import annotations

import datetime as dt
from io import BytesIO
from collections.abc import Callable, Mapping, Sequence
from typing import Any

from openpyxl import Workbook, load_workbook


class RegistryWorkbookError(ValueError):
    """Raised when a registry workbook does not match expected contract."""


# ── Enum translation maps (DB → Español para export) ────────────────────

GENDER_TRANSLATION: dict[str, str] = {
    "MALE": "MASCULINO",
    "FEMALE": "FEMENINO",
}

LICENSE_LEVEL_TRANSLATION: dict[str, str] = {
    "NATIONAL": "NACIONAL",
    "INTERNATIONAL": "INTERNACIONAL",
}

ROLE_TRANSLATION: dict[str, str] = {
    "REFEREE": "REFEREE",
    "JUDGE": "JUEZ",
    "TABLE_OFFICIAL": "OFICIAL DE MESA",
    "SUPERVISOR": "SUPERVISOR (KANSA)",
}

# Reverse maps (Español → DB para import)
GENDER_REVERSE: dict[str, str] = {v: k for k, v in GENDER_TRANSLATION.items()}
LICENSE_LEVEL_REVERSE: dict[str, str] = {
    v: k for k, v in LICENSE_LEVEL_TRANSLATION.items()
}
ROLE_REVERSE: dict[str, str] = {v: k for k, v in ROLE_TRANSLATION.items()}


def _serialize_cell_value(value: Any) -> str:
    """Convert Python values to stable workbook cell strings."""
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, bool):
        return "Sí" if value else "No"
    return str(value).strip()


def _normalize_cell_value(value: Any) -> str:
    """Normalize openpyxl cell values to string values expected by services."""
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


CellSerializer = Callable[[Mapping[str, Any]], str]


# ponytail: flat column tuples replace RegistryColumn/RegistryWorkbookAdapter
ATHLETE_SHEET_NAME = "Atletas"
ATHLETE_COLUMNS: tuple[tuple[str, str, bool, CellSerializer | None], ...] = (
    ("name", "Nombre", True, None),
    ("email", "Correo electrónico", False, None),
    ("age", "Edad", True, None),
    (
        "gender",
        "Género",
        True,
        lambda row: GENDER_TRANSLATION.get(
            str(row.get("gender", "")), str(row.get("gender", ""))
        ),
    ),
    ("weight_kg", "Peso (kg)", False, None),
    ("belt_rank", "Grado", False, None),
    ("dojo", "Dojo", False, None),
    ("nationality", "Nacionalidad (ISO3)", False, None),
    ("license_number", "Número de licencia", False, None),
)

REFEREE_SHEET_NAME = "Árbitros"
REFEREE_COLUMNS: tuple[tuple[str, str, bool, CellSerializer | None], ...] = (
    ("name", "Nombre", True, None),
    ("license_number", "Número de licencia", True, None),
    (
        "license_level",
        "Nivel de licencia",
        True,
        lambda row: LICENSE_LEVEL_TRANSLATION.get(
            str(row.get("license_level", "")), str(row.get("license_level", ""))
        ),
    ),
    (
        "role",
        "Rol",
        True,
        lambda row: ROLE_TRANSLATION.get(
            str(row.get("role", "")), str(row.get("role", ""))
        ),
    ),
    (
        "is_available",
        "Disponible",
        False,
        lambda row: _serialize_cell_value(row.get("is_available")),
    ),
    ("dojo", "Dojo", False, None),
    ("email", "Correo electrónico", False, None),
    ("phone", "Teléfono", False, None),
)


def _col_headers(
    columns: tuple[tuple[str, str, bool, CellSerializer | None], ...],
) -> tuple[str, ...]:
    """Return ordered workbook headers from column tuples."""
    return tuple(c[1] for c in columns)


def build_registry_workbook(
    sheet_name: str,
    columns: tuple[tuple[str, str, bool, CellSerializer | None], ...],
    rows: Sequence[Mapping[str, Any]],
) -> bytes:
    """Build an .xlsx workbook for a registry."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(list(_col_headers(columns)))

    for row in rows:
        cell_values: list[str] = []
        for field_name, _, _, serializer in columns:
            if serializer is not None:
                cell_values.append(serializer(row))
            else:
                cell_values.append(_serialize_cell_value(row.get(field_name)))
        worksheet.append(cell_values)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_registry_workbook(
    workbook_bytes: bytes,
    sheet_name: str,
    columns: tuple[tuple[str, str, bool, CellSerializer | None], ...],
) -> list[dict[str, str]]:
    """Parse an .xlsx workbook into canonical service row dictionaries."""
    if not workbook_bytes:
        raise RegistryWorkbookError("XLSX workbook is empty")

    try:
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl exception details vary
        raise RegistryWorkbookError("Invalid XLSX workbook") from exc

    worksheet = (
        workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
    )

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [_normalize_cell_value(value) for value in header_row]
    if not any(headers):
        raise RegistryWorkbookError("XLSX workbook is missing headers")

    header_to_column: dict[str, tuple[str, str, bool, CellSerializer | None]] = {
        c[1]: c for c in columns
    }
    missing_headers = [c[1] for c in columns if c[2] and c[1] not in headers]
    if missing_headers:
        joined = ", ".join(missing_headers)
        raise RegistryWorkbookError(f"XLSX missing required headers: {joined}")

    parsed_rows: list[dict[str, str]] = []
    for row_values in worksheet.iter_rows(min_row=2, values_only=True):
        row_dict: dict[str, str] = {}
        for index, header in enumerate(headers):
            col = header_to_column.get(header)
            if col is None:
                continue
            cell_value = row_values[index] if index < len(row_values) else None
            row_dict[col[0]] = _normalize_cell_value(cell_value)

        if any(value for value in row_dict.values()):
            parsed_rows.append(row_dict)

    return parsed_rows


def build_athletes_workbook(rows: Sequence[Mapping[str, Any]]) -> bytes:
    """Build athlete registry workbook bytes."""
    return build_registry_workbook(ATHLETE_SHEET_NAME, ATHLETE_COLUMNS, rows)


def parse_athletes_workbook(workbook_bytes: bytes) -> list[dict[str, str]]:
    """Parse athlete workbook bytes into import rows."""
    return parse_registry_workbook(workbook_bytes, ATHLETE_SHEET_NAME, ATHLETE_COLUMNS)


def build_referees_workbook(rows: Sequence[Mapping[str, Any]]) -> bytes:
    """Build referee registry workbook bytes."""
    return build_registry_workbook(REFEREE_SHEET_NAME, REFEREE_COLUMNS, rows)


def parse_referees_workbook(workbook_bytes: bytes) -> list[dict[str, str]]:
    """Parse referee workbook bytes into import rows."""
    return parse_registry_workbook(workbook_bytes, REFEREE_SHEET_NAME, REFEREE_COLUMNS)
